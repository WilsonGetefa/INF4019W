import subprocess
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
memory_image = "/home/kaliwg/Downloads/windump/WinDump.mem" # <-- Update the path -->
volatility_path = "/home/kaliwg/volatility3/vol.py"          # <-- Update the path -->

plugins = {
    "user_info": "windows.info.Info",                 # <--Show OS & kernel details of the memory sample being analyzed.-->  
    "user_assist": "windows.registry.userassist.UserAssist",            # <--Print userassist registry keys and information.-->
    #"print_key": "windows.registry.printkey.PrintKey",                # <--Print registry key and values.-->
    #"amcache": "windows.amcache.Amcache",                          # <--Extract information on executed applications from the AmCache (deprecated).-->
    #"cmdline": "windows.cmdline.CmdLine",                          # <--Lists process command line arguments.-->
    #"cmdscan": "windows.cmdscan.CmdScan",                          # <--Looks for Windows Command History lists-->
    #"getcellroutine": "windows.registry.getcellroutine.GetCellRoutine",   # <--Reports registry hives with a hooked GetCellRoutine handler-->
    #"getsids": "windows.getsids.GetSIDs",                           # <--Print the SIDs owning each process-->
    #"hivelist": "windows.registry.hivelist.HiveList",                # <--Lists the registry hives present in a particular memory image-->
    #"malfind": "windows.malware.malfind.Malfind",                   # <--Lists process memory ranges that potentially contain injected code.-->
    #"N": "windows.netscan.NetScan",                           # <--Scans for network objects present in a particular windows memory image.-->
    #"suspicious_threads": "windows.suspicious_threads.SuspiciousThreads",      # <--Lists suspicious userland process threads (deprecated).-->
    #"printkey": "windows.registry.printkey.PrintKey",                # <--Lists the registry keys under a hive or specific key value.-->
    #"pslist": "windows.pslist.PsList",                           # <--Lists the processes present in a particular windows memory image.-->
    #"psscan": "windows.psscan.PsScan",                           # <--Scans for processes present in a particular windows memory image.-->
    #"pstree": "windows.pstree.PsTree",                           # <--Plugin for listing processes in a tree based on their parent process ID.-->
    #"sessions": "windows.sessions.Sessions",                         # <--lists Processes with Session information extracted from Environmental Variables-->
    #"svcscan": "windows.svcscan.SvcScan",                           # <--Scans for windows services.-->
    #"timeliner": "timeliner.Timeliner",                             # <--Runs all relevant plugins that provide time related information and orders the results by time.-->
    #"timers": "windows.timers.Timers",                           # <--Print kernel timers and associated module DPCs-->
}

output_folder = "forensic_outputs_excel"
os.makedirs(output_folder, exist_ok=True)

# === HELPER FUNCTION TO CREATE EXCEL WITH FORMATTING ===
def save_to_excel(text, excel_file):
    lines = text.strip().splitlines()
    if not lines:
        lines = ["No output"]

    # Split lines into columns using 2+ spaces as separator
    table = [line.split() for line in lines]

    wb = Workbook()
    ws = wb.active

    # Styles
    header_font = Font(bold=True, color="FFFFFF")  # White text
    header_fill = PatternFill("solid", fgColor="4F81BD")  # Blue background
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write table
    for r, row in enumerate(table, 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = thin_border
            if r == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill

    # Freeze header row
    ws.freeze_panes = ws['A2']

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(excel_file)

# === RUN PLUGINS ===
for name, plugin in plugins.items():
    print(f"Running plugin: {plugin}")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file = os.path.join(output_folder, f"{name}_{timestamp}.xlsx")

    cmd = ["python3", volatility_path, "-f", memory_image, plugin]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        output_text = result.stdout
        save_to_excel(output_text, excel_file)
        print(f"Excel output saved to {excel_file}\n")
    except subprocess.CalledProcessError as e:
        print(f"Plugin {plugin} failed: {e}\n")
        with open(os.path.join(output_folder, f"{name}_{timestamp}_error.txt"), "w", encoding="utf-8") as f:
            f.write(f"Error running plugin:\n{e.stderr}")

print("All plugin outputs saved successfully.")