import subprocess
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
memory_image = "/home/kaliwg/Downloads/windump/WinDump.mem" # <-- Update the path -->
volatility_path = "/home/kaliwg/volatility3/vol.py"         # <-- Update the path -->

plugins = {
    "user_info": "windows.info.Info",                 # <--Show OS & kernel details of the memory sample being analyzed.-->  
    "user_assist": "windows.registry.userassist.UserAssist",            # <--Print userassist registry keys and information.-->
    "print_key": "windows.registry.printkey.PrintKey",                # <--Print registry key and values.-->
    "amcache": "windows.amcache.Amcache",                          # <--Extract information on executed applications from the AmCache (deprecated).-->
    "cmdline": "windows.cmdline.CmdLine",                          # <--Lists process command line arguments.-->
    "cmdscan": "windows.cmdscan.CmdScan",                          # <--Looks for Windows Command History lists-->
    "getcellroutine": "windows.registry.getcellroutine.GetCellRoutine",   # <--Reports registry hives with a hooked GetCellRoutine handler-->
    "getsids": "windows.getsids.GetSIDs",                           # <--Print the SIDs owning each process-->
    "hivelist": "windows.registry.hivelist.HiveList",                # <--Lists the registry hives present in a particular memory image-->
    "malfind": "windows.malware.malfind.Malfind",                   # <--Lists process memory ranges that potentially contain injected code.-->
    "N": "windows.netscan.NetScan",                           # <--Scans for network objects present in a particular windows memory image.-->
    "suspicious_threads": "windows.suspicious_threads.SuspiciousThreads",      # <--Lists suspicious userland process threads (deprecated).-->
    "printkey": "windows.registry.printkey.PrintKey",                # <--Lists the registry keys under a hive or specific key value.-->
    "pslist": "windows.pslist.PsList",                           # <--Lists the processes present in a particular windows memory image.-->
    "psscan": "windows.psscan.PsScan",                           # <--Scans for processes present in a particular windows memory image.-->
    "pstree": "windows.pstree.PsTree",                           # <--Plugin for listing processes in a tree based on their parent process ID.-->
    "sessions": "windows.sessions.Sessions",                         # <--lists Processes with Session information extracted from Environmental Variables-->
    "svcscan": "windows.svcscan.SvcScan",                           # <--Scans for windows services.-->
    #"timeliner": "timeliner.Timeliner",                             # <--Runs all relevant plugins that provide time related information and orders the results by time.-->
    "timers": "windows.timers.Timers",                           # <--Print kernel timers and associated module DPCs-->
}

output_folder = "forensic_outputs_excel_sheet"
os.makedirs(output_folder, exist_ok=True)

# === HELPER FUNCTION TO ADD SHEET TO EXCEL WITH FORMATTING ===
def add_sheet(ws, text):
    lines = text.strip().splitlines()
    if not lines:
        lines = ["No output"]

    # Split lines into columns (simple whitespace split)
    table = [line.split() for line in lines]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")  # White text
    header_fill = PatternFill("solid", fgColor="4F81BD")  # Blue background
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write table
    for r, row in enumerate(table, 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = thin_border
            if r == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill

    # Freeze header
    ws.freeze_panes = ws['A2']

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# === CREATE SINGLE WORKBOOK ===
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
workbook_file = os.path.join(output_folder, f"forensic_report_{timestamp}.xlsx")
wb = Workbook()
# Remove the default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# === RUN PLUGINS AND ADD SHEETS ===
for name, plugin in plugins.items():
    print(f"Running plugin: {plugin}")
    ws = wb.create_sheet(title=name[:31])  # Excel sheet name max 31 chars

    cmd = ["python3", volatility_path, "-f", memory_image, plugin]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        output_text = result.stdout
        add_sheet(ws, output_text)
        print(f"Added sheet for {name}")
    except subprocess.CalledProcessError as e:
        print(f"Plugin {plugin} failed: {e}")
        ws.cell(row=1, column=1, value=f"Error running plugin:\n{e.stderr}")

# === SAVE WORKBOOK ===
wb.save(workbook_file)
print(f"All plugin outputs saved in single Excel workbook: {workbook_file}")